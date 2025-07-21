import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Membuat workbook baru
wb = Workbook()

# Sheet 1: Kriteria dan Bobot
ws1 = wb.active
ws1.title = 'Kriteria AHP'

# Styling
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Judul utama
ws1['A1'] = 'SIMULASI PERHITUNGAN AHP'
ws1['A2'] = 'SISTEM EVALUASI GURU SMP PENIDA KATAPANG'
ws1['A3'] = 'ANALISIS KRITERIA EVALUASI'

# Merge cells untuk judul
ws1.merge_cells('A1:H1')
ws1.merge_cells('A2:H2')
ws1.merge_cells('A3:H3')

# Style judul
for row in range(1, 4):
    cell = ws1[f'A{row}']
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')

# Data Kriteria Evaluasi
kriteria_data = [
    ['K1', 'Kedisiplinan', 'Ketepatan waktu masuk, absensi, kepatuhan aturan', 0.25],
    ['K2', 'Penguasaan Materi', 'Kemampuan menguasai mata pelajaran yang diampu', 0.30],
    ['K3', 'Metode Mengajar', 'Variasi metode, penggunaan media, interaksi siswa', 0.20],
    ['K4', 'Komunikasi', 'Kemampuan berkomunikasi dengan siswa dan rekan', 0.15],
    ['K5', 'Evaluasi Pembelajaran', 'Sistem penilaian, feedback, remedial', 0.10]
]

# Header tabel kriteria
headers_kriteria = ['Kode', 'Kriteria', 'Deskripsi', 'Bobot AHP']
row_start = 5

for col, header in enumerate(headers_kriteria, 1):
    cell = ws1.cell(row=row_start, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Input data kriteria
for i, row_data in enumerate(kriteria_data, row_start + 1):
    for j, value in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=j, value=value)
        cell.border = border
        if j == 4:  # Format persentase untuk bobot
            cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center' if j <= 2 else 'left', vertical='center')

# Adjust column widths
column_widths = [8, 20, 50, 12]
for i, width in enumerate(column_widths, 1):
    ws1.column_dimensions[chr(64 + i)].width = width

# Sheet 2: Matriks Perbandingan Berpasangan
ws2 = wb.create_sheet(title='Matriks Perbandingan')

# Judul untuk sheet 2
ws2['A1'] = 'MATRIKS PERBANDINGAN BERPASANGAN KRITERIA'
ws2.merge_cells('A1:G1')
ws2['A1'].font = Font(bold=True, size=12)
ws2['A1'].alignment = Alignment(horizontal='center')

# Header matriks perbandingan
kriteria_codes = ['', 'K1', 'K2', 'K3', 'K4', 'K5']
row_start_matrix = 3

for col, code in enumerate(kriteria_codes, 1):
    cell = ws2.cell(row=row_start_matrix, column=col, value=code)
    if col == 1:
        cell.font = Font(bold=True)
    else:
        cell.font = header_font
        cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Data matriks perbandingan (contoh)
matrix_data = [
    ['K1', 1, 0.5, 2, 3, 4],
    ['K2', 2, 1, 3, 4, 5],
    ['K3', 0.5, 0.33, 1, 2, 3],
    ['K4', 0.33, 0.25, 0.5, 1, 2],
    ['K5', 0.25, 0.2, 0.33, 0.5, 1]
]

for i, row_data in enumerate(matrix_data, row_start_matrix + 1):
    for j, value in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=value)
        cell.border = border
        if j == 1:
            cell.font = Font(bold=True)
            cell.fill = subheader_fill
        else:
            cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Sheet 3: Hasil Evaluasi Guru
ws3 = wb.create_sheet(title='Hasil Evaluasi')

# Judul untuk sheet 3
ws3['A1'] = 'HASIL EVALUASI GURU'
ws3['A2'] = 'BERDASARKAN METODE AHP'
ws3.merge_cells('A1:I1')
ws3.merge_cells('A2:I2')

for row in range(1, 3):
    cell = ws3[f'A{row}']
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

# Data guru contoh - GUNAKAN INI (nilai manual dengan format yang benar)
guru_data = [
    [1, 'Ahmad Fauzi, S.Pd', 'Matematika', 85, 90, 80, 75, 85, 84.25],
    [2, 'Siti Nurhaliza, S.Pd', 'Bahasa Indonesia', 90, 85, 85, 80, 90, 86.0],
    [3, 'Budi Santoso, S.Pd', 'IPA', 80, 95, 90, 85, 80, 87.0],
    [4, 'Rina Wati, S.Pd', 'Bahasa Inggris', 95, 80, 75, 90, 85, 84.25],
    [5, 'Dedi Kurnia, S.Pd', 'IPS', 75, 85, 95, 80, 90, 83.25]
]

# Header tabel evaluasi
headers_evaluasi = ['No', 'Nama Guru', 'Mata Pelajaran', 'Kedisiplinan', 'Penguasaan Materi', 'Metode Mengajar', 'Komunikasi', 'Evaluasi Pembelajaran', 'Nilai AHP']
row_start_eval = 4

for col, header in enumerate(headers_evaluasi, 1):
    cell = ws3.cell(row=row_start_eval, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Input data evaluasi guru dengan format yang benar
for i, row_data in enumerate(guru_data, row_start_eval + 1):
    for j, value in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=j, value=value)
        cell.border = border
        if j > 3 and j <= 8:  # Format angka untuk nilai kriteria
            cell.number_format = '0.00'  # Format dengan 2 desimal
        elif j == 9:  # Format untuk nilai AHP
            cell.number_format = '0.00'  # Format dengan 2 desimal
        cell.alignment = Alignment(horizontal='center' if j <= 3 else 'right', vertical='center')

# Adjust column widths untuk sheet 3
column_widths_eval = [5, 20, 15, 12, 15, 12, 12, 18, 12]
for i, width in enumerate(column_widths_eval, 1):
    ws3.column_dimensions[chr(64 + i)].width = width

# Sheet 4: Ranking dan Rekomendasi
ws4 = wb.create_sheet(title='Ranking')

# Judul untuk sheet 4
ws4['A1'] = 'RANKING GURU BERDASARKAN AHP'
ws4['A2'] = 'DAN REKOMENDASI PERBAIKAN'
ws4.merge_cells('A1:F1')
ws4.merge_cells('A2:F2')

for row in range(1, 3):
    cell = ws4[f'A{row}']
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')

# Header ranking
headers_ranking = ['Ranking', 'Nama Guru', 'Nilai AHP', 'Kategori', 'Rekomendasi']
row_start_rank = 4

for col, header in enumerate(headers_ranking, 1):
    cell = ws4.cell(row=row_start_rank, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Data ranking - sesuaikan dengan nilai yang benar
ranking_data = [
    [1, 'Budi Santoso, S.Pd', 87.0, 'Sangat Baik', 'Pertahankan kinerja, jadikan mentor'],
    [2, 'Siti Nurhaliza, S.Pd', 86.0, 'Baik', 'Variasikan metode mengajar'],
    [3, 'Ahmad Fauzi, S.Pd', 84.25, 'Baik', 'Perbaiki komunikasi dengan siswa'],
    [4, 'Rina Wati, S.Pd', 84.25, 'Baik', 'Tingkatkan penguasaan materi'],
    [5, 'Dedi Kurnia, S.Pd', 83.25, 'Cukup', 'Tingkatkan kedisiplinan dan evaluasi']
]

# Input data ranking
for i, row_data in enumerate(ranking_data, row_start_rank + 1):
    for j, value in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=j, value=value)
        cell.border = border
        if j == 3:  # Format angka untuk nilai AHP
            cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center' if j <= 4 else 'left', vertical='center')

# Adjust column widths untuk sheet 4
column_widths_rank = [10, 20, 12, 15, 40]
for i, width in enumerate(column_widths_rank, 1):
    ws4.column_dimensions[chr(64 + i)].width = width

# Menambahkan keterangan AHP di sheet pertama
ws1['A' + str(len(kriteria_data) + row_start + 3)] = 'Keterangan AHP:'
ws1['A' + str(len(kriteria_data) + row_start + 4)] = 'AHP = Analytical Hierarchy Process'
ws1['A' + str(len(kriteria_data) + row_start + 5)] = 'Metode pengambilan keputusan dengan perbandingan berpasangan'
ws1['A' + str(len(kriteria_data) + row_start + 6)] = 'Consistency Ratio (CR) harus < 0.1 untuk validitas hasil'

# Style keterangan
for row in range(len(kriteria_data) + row_start + 3, len(kriteria_data) + row_start + 7):
    ws1[f'A{row}'].font = Font(italic=True)

# Save file
wb.save('/Users/flashcode/Documents/project-destra/AHP_Evaluasi_Guru_SMP_PENIDA_KATAPANG_NEW.xlsx')
print('File Excel AHP Evaluasi Guru berhasil dibuat!')
print('File tersimpan di: /Users/flashcode/Documents/project-destra/AHP_Evaluasi_Guru_SMP_PENIDA_KATAPANG_NEW.xlsx')